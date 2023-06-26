import csv, re, datetime, openpyxl, os, locale
from openpyxl.utils import get_column_letter

TULEMUSTEFAILINIMI = "tulemused.xlsx"

def küsi_failinimi(küsimus, muster = ""):
    """
    Küsib kasutajalt failinime või võtab selle etteantud mustri järgi ise, kui leiab.
    """
    if muster:
        failinimed = [failinimi for failinimi in os.listdir() if re.search(muster, failinimi)]
        if len(failinimed) == 0:
            print("Jooksvast kaustast sobivaid faile ei leidnud.")
        elif len(failinimed) == 1:
            print(f"Leidsin faili '{failinimed[0]}'.")
            return failinimed[0]
        else:
            print(f"Leidsin {len(failinimed)} sobivat faili:")
            print("\n".join(failinimed))

    while True:
        failinimi = input(küsimus.rstrip()+" ").strip()
        if not failinimi:
            return
        if os.path.exists(failinimi):
            return failinimi
        print("Sellist faili pole.")

def loe_failid(punktidenimi, vastustenimi):
    """
    Loeb punktide failist ja vastuste failist küsimuste tekstid, vastused ja punktid.

    Tagastab:
        list: Kõigi andmete tabel
    """
    punktid = []
    with open(punktidenimi, newline='', encoding="utf8") as f:
        lugeja = csv.reader(f)
        for rida in lugeja:
            if "Perenimi" in rida[0] or "Surname" in rida[0]:
                maksimumid = [el.split("/")[1] for el in rida[6:]]
                continue
            if "Üldine keskmine" in rida[0] or "Overall average" in rida[0]:
                break
            punktid.append(rida)

    vastused = []
    with open(vastustenimi, newline='', encoding="utf8") as f:
        lugeja = csv.reader(f)
        for rida in lugeja:
            if "Perenimi" in rida[0] or "Surname" in rida[0]:
                algindeks = rida.index("Küsimus 1") if "Küsimus 1" in rida else rida.index("Question 1")
                ülarv = (len(rida) - algindeks) // 2
                print("Küsimuste arv", ülarv)
                continue
            vastused.append(rida)

    if len(punktid) != len(vastused):
        print("Punktide ja vastuste faili ridade arvud on erinevad.")
        return []

    locale.setlocale(locale.LC_ALL, "")
    punktid.sort(key=lambda x: (locale.strxfrm(x[0]), locale.strxfrm(x[1]), x[3]))
    vastused.sort(key=lambda x: (locale.strxfrm(x[0]), locale.strxfrm(x[1]), x[3]))

    for i in range(len(punktid)):
        if punktid[i][:3] != vastused[i][:3]:
            print("Punktide ja vastuste failis olevad nimed ei lange kokku.")
            return []

    tabel = []
    for prida, vrida in zip(punktid, vastused):
        # Eesnimi, perekonnanimi, algusaeg, lõppaeg
        tabelirida = [prida[1], prida[0], prida[3], prida[4]]

        # Küsimus, vastus, punktid
        for j in range(ülarv):
            indeks = -2 * (ülarv - j)
            küsimus = vrida[indeks].replace('\\"', '"').replace('\\\\n', '\\n')
            vastus = vrida[indeks+1].replace('\\"', '"').replace('\\\\n', '\\n')
            punktid = prida[j+7] + '/' + maksimumid[j+1] if len(prida) > j+7 else '-'
            tabelirida.extend([küsimus, vastus, punktid])

        # Kogusumma
        tabelirida.append((prida[6] + '/' + maksimumid[0] if len(prida) > 6 else '-'))
        tabel.append(tabelirida)

    return tabel

def teisenda_kuupäevad(tabel, n):
    """
    Teisendab tabelis kuupäevade veeru klassi datetime objektideks.
    """
    for rida in tabel:
        # Kontrolli, kas veerg on sobiva kujuga
        if not re.search(r'\d+[.]\d+[.]\d+\s+\d+:\d+:\d+', rida[n]):
            continue

        # Eralda päev, kuu, aasta, tund, minut, sekund
        s = rida[n].split()
        päev, kuu, aasta = map(int, s[0].split('.'))
        tund, minut, sekund = map(int, s[1].split(':'))
        rida[n] = datetime.datetime(aasta, kuu, päev, tund, minut, sekund)

def teisenda_küsimused(tabel, n):
    """
    Muudab küsimuste veerus olevad unikaalsed väärtused järjestikusteks arvudeks.
    """

    # Kogu erinevad küsimused kokku
    küsimused = set()
    for rida in tabel:
        küsimus = rida[n]

        # Sorteeri valikvastustega küsimuse vastusevariandid
        m = re.search(r': ((?:.*; )+.*)$', küsimus)
        if m:
            vastused = '; '.join(sorted(m.group(1).split("; ")))
            küsimus = küsimus[:m.start(1)] + vastused
            rida[n] = küsimus

        # Lünkteksti valikvastustega küsimused
        alamosad = re.findall(r'\{(?:[^\n]*; )+[^\n]*\}', küsimus)
        for alamosa in alamosad:
            variandid = sorted(map(str.strip, alamosa[1:-1].split('; ')))
            uusosa = '{' + '; '.join(variandid) + '}'
            küsimus = küsimus.replace(alamosa, uusosa)
        rida[n] = küsimus

        # Pildile lohistamise küsimused
        m = re.search(r'; \[\[.*\]\] -> \{.*\}.*', küsimus)
        if m:
            küsimus = küsimus[:m.start(0)]
            rida[n] = küsimus

        küsimused.add(küsimus)

    küsimused = sorted(küsimused)

    # Määra pikkadele küsimustele numbrid
    teisendus = {}
    for i, küsimus in enumerate(küsimused, start=1):
        if len(küsimus) > 10:
            teisendus[küsimus] = i

    # Asenda küsimused numbritega
    if len(teisendus) > 0:
        for rida in tabel:
            küsimus = rida[n]
            if küsimus in teisendus:
                rida[n] = teisendus[küsimus]


TÜÜP_TEADMATA = 0
TÜÜP_STACK = 1
TÜÜP_LÜNKTEKST = 2
TÜÜP_LOHISTA_PILDILE = 3

def küsimuse_tüüp(s):
    """
    Tagastab sõne väärtuse järgi küsimuse tüübi.
    """
    if s.startswith(('Seed:', 'ans1:', 'prt1:')):
        return TÜÜP_STACK
    elif s.startswith(('osa 1:', 'part 1:', 'часть 1:', 'partie 1 :', 'Teil 1:')):
        return TÜÜP_LÜNKTEKST
    elif s.startswith(('Kukutusala 1', 'Drop zone 1', 'Зона 1', 'Zone de dépôt 1', 'Dropzone 1')):
        return TÜÜP_LOHISTA_PILDILE
    else:
        return TÜÜP_TEADMATA

def teisenda_vastused(tabel, n):
    """
    Eraldab vastuste veeru lahtritest välja tegelikud vastused.
    Kirjutab tegelike vastuste järjendi samasse lahtrisse tagasi.
    """
    for rida in tabel:
        s = rida[n]
        tüüp = küsimuse_tüüp(s)

        if tüüp == TÜÜP_STACK:
            if not re.search("ans[0-9]+: ", s):
                rida[n] = '-'
                continue
            osad = re.findall(r'ans[0-9]+: "?(.*?)"? \[(?:score|invalid)\];', s)
        elif tüüp == TÜÜP_LÜNKTEKST:
            osad = re.findall(r'(?:osa [0-9]+|part [0-9]+|часть [0-9]+|partie [0-9]+ |Teil [0-9]+): '
                              r'(.*?)(?:; |$)', s)
        elif tüüp == TÜÜP_LOHISTA_PILDILE:
            osad = re.findall(r'{.*? ([0-9]+)}', s)
        elif '; ' in s:
            osad = sorted(s.split('; '))
        else:
            osad = s

        rida[n] = osad if osad else ''

def teisenda_punktid(tabel, n):
    """
    Teisendab tabeli veerus punktid ujukomaarvudeks ja ümardab sobivalt.
    """
    for rida in tabel:
        if '/' in rida[n]:
            try:
                a, m = rida[n].split('/', maxsplit=1)
                kohtade_arv = len(m.split('.')[-1]) if '.' in m else 0
                a, m = map(float, (a, m))
                astmete_arv = round(m * 10**kohtade_arv)
                if astmete_arv <= 1:
                    k = 1
                elif astmete_arv <= 2:
                    k = 0.5
                elif astmete_arv <= 5:
                    k = 0.2
                elif astmete_arv <= 10:
                    k = 0.1
                elif astmete_arv <= 20:
                    k = 0.05
                elif astmete_arv <= 50:
                    k = 0.02
                else:
                    k = 0.01
                x = round(a / m / k) * k
                rida[n] = round(x, 2)
            except:
                rida[n] = '-'

def lisa_märked(tabel):
    """
    Lisab tabelisse koodide veeru: kas katse on esimene ja kas katse on parim.
    """
    sõn = {}
    for i, rida in enumerate(tabel):
        nimi = rida[0] + " " + rida[1]
        if nimi in sõn:
            if isinstance(rida[-1], float) and isinstance(tabel[sõn[nimi]['parim']][-1], float) and \
                    rida[-1] > tabel[sõn[nimi]['parim']][-1]:
                sõn[nimi]['parim'] = i
        else:
            sõn[nimi] = {'esimene': i, 'parim': i}

    for i, rida in enumerate(tabel):
        nimi = rida[0] + " " + rida[1]
        if i == sõn[nimi]['esimene'] and i == sõn[nimi]['parim']:
            rida.append('11')
        elif i == sõn[nimi]['esimene'] and i != sõn[nimi]['parim']:
            rida.append('10')
        elif i != sõn[nimi]['esimene'] and i == sõn[nimi]['parim']:
            rida.append('01')
        else:
            rida.append('00')

def korrasta_tabel(tabel):
    """
    Korrastab tabeli veerud.
    Eesnimi, Perenimi, Alustatud, Lõpetatud, {Küsimus, Vastus, Punktid}*, Kokku, Märked
    """
    teisenda_kuupäevad(tabel, 2)
    teisenda_kuupäevad(tabel, 3)

    for n in range(4, len(tabel[0]) - 1, 3):
        teisenda_küsimused(tabel, n)

    for n in range(5, len(tabel[0]) - 1, 3):
        teisenda_vastused(tabel, n)

    for n in range(6, len(tabel[0]) - 1, 3):
        teisenda_punktid(tabel, n)

    teisenda_punktid(tabel, -1)
    lisa_märked(tabel)

def kirjuta_fail(tabel):
    """
    Salvestab tabeli Exceli faili.
    """
    wb = openpyxl.Workbook()
    ülarv = (len(tabel[0]) - 6) // 3
    for i in range(ülarv):
        if i == 0:
            leht = wb.active
            leht.title = "1"
        else:
            leht = wb.create_sheet(f"{i+1}")

        # Päiserida
        leht['A1'] = 'Nr'
        leht['B1'] = 'Eesnimi'
        leht['C1'] = 'Perenimi'
        leht['D1'] = 'EP'
        leht['E1'] = 'Küs'

        veerunr = 3*i+5
        vastustearv = max(len(rida[veerunr]) if isinstance(rida[veerunr], list) else 1 for rida in tabel)
        for j in range(vastustearv):
            leht.cell(row=1, column=6+j, value=f'Vas {j+1}')
        leht.cell(row=1, column=6+vastustearv, value='Pun')
        leht.cell(row=1, column=7+vastustearv, value='Uus')

        # Sisuread
        sisuread = []
        for m, rida in enumerate(tabel):
            uusrida = [m+1, rida[0], rida[1], rida[-1], rida[veerunr-1]]
            vastus = rida[veerunr]
            for j in range(vastustearv):
                if isinstance(vastus, list):
                    uusrida.append(vastus[j] if j < len(vastus) else "")
                else:
                    uusrida.append(vastus if j == 0 else "")
            uusrida.append(rida[veerunr+1])
            sisuread.append(uusrida)

        for j in range(4 + vastustearv, 4, -1):
            sisuread.sort(key=lambda x: ' ' + x[j] if x[j] != '-' else x[j])
        sisuread.sort(key=lambda x: x[5+vastustearv]
                      if isinstance(x[5+vastustearv], float) else -1, reverse=True)

        for m, rida in enumerate(sisuread):
            for j in range(6 + vastustearv):
                if isinstance(rida[j], str) and rida[j].startswith('='):
                    rida[j] = " " + rida[j]
                leht.cell(row=m+2, column=1+j, value=rida[j])
            for veerg in ['D', 'E']:
                leht[f'{veerg}{m+2}'].alignment = openpyxl.styles.Alignment(horizontal='center')

        # Kujundus
        for j in range(vastustearv + 7):
            leht.cell(row=1, column=1+j).font = openpyxl.styles.Font(bold=True)
        leht.column_dimensions['A'].width = 4  # Nr
        leht.column_dimensions['D'].width = 5  # EP
        leht.column_dimensions['E'].width = 5  # Küs
        for j in range(vastustearv):
            maxlaius = 0
            for m in range(len(tabel)):
                try:
                    laius = len(leht[get_column_letter(6 + j) + str(m + 2)].value)
                except:
                    laius = 0
                if laius > maxlaius:
                    maxlaius = laius
            leht.column_dimensions[get_column_letter(6 + j)].width = min(maxlaius + 2, 30)
        leht.column_dimensions[get_column_letter(6 + vastustearv)].width = 5  # Pun
        leht.column_dimensions[get_column_letter(7 + vastustearv)].width = 5  # Uus

        # Filter
        leht.auto_filter.ref = 'A1:' + get_column_letter(7 + vastustearv) + str(len(tabel) + 1)

    wb.save(TULEMUSTEFAILINIMI)

def main():
    """
    Küsib punktide ja vastuste failide nimesid, loeb vastused sisse ja moodustab Exceli faili.
    """
    punktidenimi = küsi_failinimi("Sisesta punktide faili nimi:", r'-(hinded|grades)( [(][0-9]+[)])?[.]csv$')
    if not punktidenimi:
        return

    vastustenimi = küsi_failinimi("Sisesta vastuste faili nimi:", r'-(vastused|responses)( [(][0-9]+[)])?[.]csv$')
    if not vastustenimi:
        return

    tabel = loe_failid(punktidenimi, vastustenimi)

    if tabel:
        korrasta_tabel(tabel)
        kirjuta_fail(tabel)

if __name__ == "__main__":
    main()