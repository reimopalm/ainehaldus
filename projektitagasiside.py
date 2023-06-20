import openpyxl, re, os.path, statistics


def veerutäht(j):
    if 1 <= j <= 26:
        return chr(64 + j)
    else:
        return chr(64 + (j - 1) // 26) + chr(65 + (j - 1) % 26)


def lahtritähis(i, j):
    return veerutäht(j) + str(i)


# i on reanumber, j on veerunumber
# Ülemine vasak nurk on 1, 1
def lahtrisisu(leht, i, j):
    sisu = leht[lahtritähis(i, j)].value
    if sisu == None:
        return ''
    return sisu


def loeandmed(failinimi):
    # Loeb failist andmed
    # Tagastab projektide sõnastiku ja hindajate sõnastiku
    projektis_veerge = 5
    töövihik = openpyxl.load_workbook(failinimi)
    leht = töövihik.worksheets[0]
    ridade_arv = leht.max_row
    veergude_arv = leht.max_column
    projektide_arv = (veergude_arv - 3) // projektis_veerge

    projektid = {}
    hindajad = {}

    for i in range(2, ridade_arv + 1):
        hindaja_nimi = lahtrisisu(leht, i, 1).strip()
        hindaja_rühm = lahtrisisu(leht, i, 2).strip()
        if hindaja_nimi == '':
            continue
        if hindaja_nimi not in hindajad:
            hindaja = dict()
            hindaja['rühm'] = hindaja_rühm
            hindaja['väärtus'] = []
            hindaja['meeldimised'] = []
            hindaja['soovitused'] = []
            hindaja['esitlusehinded'] = []
            hindajad[hindaja_nimi] = hindaja

        for n in range(projektide_arv):
            projektinimi = lahtrisisu(leht, i, 5 * n + 4).strip()
            if projektinimi == '':
                continue
            progväärtus = lahtrisisu(leht, i, 5 * n + 5).strip()
            meeldis = lahtrisisu(leht, i, 5 * n + 6).strip()
            soovitus = lahtrisisu(leht, i, 5 * n + 7).strip()
            esitlusehinne = lahtrisisu(leht, i, 5 * n + 8).strip()

            if projektinimi not in projektid:
                projektid[projektinimi] = dict()
                projektid[projektinimi]['väärtus'] = []
                projektid[projektinimi]['meeldimised'] = []
                projektid[projektinimi]['soovitused'] = []
                projektid[projektinimi]['esitlusehinded'] = []

            if progväärtus != '' and hindaja_rühm not in projektinimi:
                # Enda projekti väärtuse hindamine jätta vahele
                projektid[projektinimi]['väärtus'].append(progväärtus)
                hindajad[hindaja_nimi]['väärtus'].append(progväärtus)
            if meeldis != '':
                # Enda projekti meeldimised võtta sisse
                projektid[projektinimi]['meeldimised'].append(meeldis)
                hindajad[hindaja_nimi]['meeldimised'].append(meeldis)
            if soovitus != '':
                # Enda projekti soovitused võtta sisse
                projektid[projektinimi]['soovitused'].append(soovitus)
                hindajad[hindaja_nimi]['soovitused'].append(soovitus)
            if esitlusehinne != '' and hindaja_rühm not in projektinimi:
                # Enda projekti esitluse hindamine jätta vahele
                projektid[projektinimi]['esitlusehinded'].append(esitlusehinne)
                hindajad[hindaja_nimi]['esitlusehinded'].append(esitlusehinne)
    return projektid, hindajad


def väljastaprojektid(projektid, algnimi):
    # Väljastab projektide tagasiside html-faili
    # * Punkte projekti väärtusele - keskmine, mediaan
    # * Mis meeldis - kommentaarid
    # * Mis ei meeldinud - kommentaarid
    # * Punkte esitlusele - hinnangud, keskmine

    baasnimi = os.path.splitext(algnimi)[0]
    failinimi = baasnimi + '-proj.html'
    projnimed = sorted(projektid.keys())
    f = open(failinimi, 'w')

    for projnimi in projnimed:
        f.write('<h1>Project ' + projnimi + '</h1>\n')

        a = projektid[projnimi]['väärtus']
        pealkirjad = [
            'Program has exceptional value, can be used in various situations, and provides substantial help to the user.',
            'Program has good value, fills its place, and provides good help to the user.',
            'Program is useful in limited situations or with a limited number of inputs.',
            'Program carries little or no added value; its objectives can easily be achieved by other means.']
        f.write('<p><b>How do you rate the general value of the program?</b></p>\n')
        f.write('<ul>\n')
        summa = kokku = 0
        arvud = []
        for i in range(len(pealkirjad)):
            x = pealkirjad[i]
            esinemistearv = a.count(x)
            if esinemistearv > 0:
                f.write(f'<li>{x.rstrip(".")}: {a.count(x)}</li>\n')
                summa += (3 - i) * esinemistearv
                kokku += esinemistearv
                arvud += [3 - i] * esinemistearv
        f.write('</ul>\n')
        keskmine = str(round(summa / kokku, 1)).rstrip('.0')
        mediaan = str(round(statistics.median(arvud), 1)).rstrip('.0')
        f.write(f'<p>Average: {keskmine}/3<br>\n')
        f.write(f'Median: {mediaan}/3</p>\n')

        f.write('<p><b>What did you like in this project?</b></p>\n')
        f.write('<ul>\n')
        for rida in projektid[projnimi]['meeldimised']:
            puhasrida = rida.replace('\n', '<br>')
            while '&amp;' in puhasrida: puhasrida = puhasrida.replace('&amp;', '&')
            f.write('<li>' + puhasrida + '</li>\n')
        f.write('</ul>\n')

        f.write('<p><b>What could have been done differently?</b></p>\n')
        f.write('<ul>\n')
        for rida in projektid[projnimi]['soovitused']:
            if '&amp' in rida:
                print(repr(rida))
            puhasrida = rida.replace('\n', '<br>')
            while '&amp;' in puhasrida: puhasrida = puhasrida.replace('&amp;', '&')
            f.write('<li>' + puhasrida + '</li>\n')
        f.write('</ul>\n')

        a = projektid[projnimi]['esitlusehinded']
        pealkirjad = ['Presentation is informative, correct, complete, and convincing.',
                      'Presentation is mostly informative and mostly complete.',
                      'Presentation is somewhat informative but incomplete or unconvincing.',
                      'Presentation is incoherent, incomprehensible, or incorrect.']
        f.write('<p><b>How do you rate the presentation?</b></p>\n')
        f.write('<ul>\n')
        summa = kokku = 0
        arvud = []
        for i in range(len(pealkirjad)):
            x = pealkirjad[i]
            esinemistearv = a.count(x)
            if esinemistearv > 0:
                f.write(f'<li>{x.rstrip(".")}: {a.count(x)}</li>\n')
                summa += (3 - i) * esinemistearv
                kokku += esinemistearv
                arvud += [3 - i] * esinemistearv
        f.write('</ul>\n')
        keskmine = str(round(summa / kokku, 1)).rstrip('.0')
        mediaan = str(round(statistics.median(arvud), 1)).rstrip('.0')
        f.write(f'<p>Average: {keskmine}/3<br>\n')
        f.write(f'Median: {mediaan}/3</p>\n')

    f.close()


def väljastahindajad(hindajad, algnimi):
    # Väljastab hindaja tegevuste kokkuvõtte.
    # * Mis meeldis - vastuste arv
    # * Mis ei meeldinud - vastuste arv
    # * Punkte projektile - vastuste arv
    # * Punkte esitlusele - vastuste arv
    baasnimi = os.path.splitext(algnimi)[0]
    failinimi = baasnimi + '-hind.txt'
    hindajanimed = sorted(hindajad.keys())
    f = open(failinimi, 'w')
    for hindajanimi in hindajanimed:
        progväärtusarv = len(hindajad[hindajanimi]['väärtus'])
        meeldimistearv = len(hindajad[hindajanimi]['meeldimised'])
        soovitustearv = len(hindajad[hindajanimi]['soovitused'])
        esithindearv = len(hindajad[hindajanimi]['esitlusehinded'])
        f.write(f'{hindajanimi}\t{progväärtusarv}\t{meeldimistearv}\t{soovitustearv}\t{esithindearv}\n')
    f.close()


failinimi = 'Feedback to the presentations.xlsx'
projektid, hindajad = loeandmed(failinimi)
väljastaprojektid(projektid, failinimi)
väljastahindajad(hindajad, failinimi)